#print(sheet["B3"].value)  # .value позволяет видеть значение указанной ячейки, без него - только название листа Excel
# ЗНО0310189066
# Также можно обратиться к ячейке по номеру строки (начинаются с 1) и номеру столбца (начинаются с 0):
#print(sheet[3][1].value)
# ЗНО0310189066

# Если нужно обработать данные в диапазоне таблицы, указываем его в переменной, н-р, cells = sheet['B1':'C5']
# # Обход данного диапазона выведет кортеж:

# # for zno, status in cells:
# #     print(zno.value, status.value) -- указываем только объявленные выше переменные в любом порядке

# # Либо использовать встроенный метод iter_rows, указывая мин и макс значения строк и столбцов - тут нумерация всего обычная!:
# """ import openpyxl
# sheet = book.active
# book = openpyxl.open("C:/Users/Noizz/Desktop/learn/Excel отчёт/Сводка по объектам SM.xlsx", read_only=True)

# for row in sheet.iter_rows(min_row=2, max_row=20, min_col=2, max_col=3):
# # ЕСЛИ НЕ ПЕРЕДОВАТЬ МИН И МАКС ЗНАЧЕНИЯ - ПРОЙДЁМ ПО ВСЕЙ ТАБЛИЦЕ: for row in sheet.iter_rows():
#     for cell in row:
#         print(cell.value, sep = ' ')
#     print() """

# # Для обращения к другому листу Excel используется атрибут worksheets - это список:

# import openpyxl

# book = openpyxl.open("C:/Users/Noizz/Desktop/learn/Excel отчёт/Сводка по объектам SM.xlsx", read_only=True)

# sheet_2 = book.worksheets[2]
# print(sheet_2["B4"].value)


import openpyxl
book = openpyxl.open("C:/Users/Noizz/Desktop/learn/Excel отчёт/Сводка по объектам SM.xlsx", read_only=True)

sheet = book.active

for row in range(1, sheet.max_row + 1):
    zno = sheet[row][1].value
    object = sheet[row][5].value
    control_time = sheet[row][6].value
    group = sheet[row][8].value
    employee = sheet[row][9].value
    print(zno, object, control_time, group, employee)
    print()

sheet_1 = book.worksheets[1]

for row in range(1, sheet_1.max_row + 1):
    znr = sheet_1[row][1].value
    object = sheet_1[row][3].value
    control_time = sheet_1[row][4].value
    group = sheet_1[row][5].value
    employee = sheet_1[row][6].value
    print(znr, object, control_time, group, employee)

sheet_2 = book.worksheets[2]

for row in range(1, sheet_2.max_row + 1):
    incident = sheet_2[row][1].value
    object = sheet_2[row][4].value
    control_time = sheet_2[row][5].value
    group = sheet_2[row][7].value
    employee = sheet_2[row][8].value
    print(incident, object, control_time, group, employee)

sheet_4 = book.worksheets[4]

for row in range(1, sheet_2.max_row + 1):
    mpr = sheet_4[row][1].value
    object = sheet_4[row][4].value
    control_time = sheet_4[row][5].value
    group = sheet_4[row][8].value
    employee = sheet_4[row][7].value
    print(mpr, object, control_time, group, employee)
