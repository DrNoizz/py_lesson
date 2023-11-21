import openpyxl, pytz
# from openpyxl.styles import Font, PatternFill
from datetime import datetime, time

book = openpyxl.open("//samora101.sigma.sbrf.ru/VOL1_FolderRedirection/STD-VARM/11901848/My Documents/py/Сводка по объектам SM.xlsx")
sheet = book.active

for row in range(2, sheet.max_row + 1):  
    zno = sheet[row][1].value
    object = sheet[row][5].value
    control_time = sheet[row][6].value
    group = sheet[row][2].value
    employee = sheet[row][8].value
    date = control_time
    date = date.replace(tzinfo=pytz.timezone('Europe/Moscow')) 
    # format = "%y/%m/%d %H:%M:%S"

    dt = datetime.now(pytz.timezone('Europe/Moscow'))
    dt = dt.replace(microsecond=0) 

    if date < dt:
        # sheet.delete_rows(2)
        print(zno, object, control_time, group, employee, sep='\t')
        

        
sheet = book['Инциденты']

for row in range(2, sheet.max_row + 1):  
    incident = sheet[row][1].value
    object = sheet[row][4].value
    control_time = sheet[row][5].value
    group = sheet[row][7].value
    employee = sheet[row][7].value
    date = control_time
    date = date.replace(tzinfo=pytz.timezone('Europe/Moscow'))
    # format = "%y/%m/%d %H:%M:%S"

    dt = datetime.now(pytz.timezone('Europe/Moscow'))
    dt = dt.replace(microsecond=0) 

    if date < dt:
        # sheet.delete_rows(2)
        print(incident, object, control_time, group, employee, sep='\t')   
        
sheet = book['Мероприятия по рискам']

for row in range(2, sheet.max_row + 1):  
    mpr = sheet[row][1].value
    object = sheet[row][4].value
    control_time = sheet[row][5].value
    group = sheet[row][8].value
    employee = sheet[row][7].value
    date = control_time
    date = date.replace(tzinfo=pytz.timezone('Europe/Moscow')) 
    # format = "%y/%m/%d %H:%M:%S"

    dt = datetime.now(pytz.timezone('Europe/Moscow'))
    dt = dt.replace(microsecond=0) 

    if date < dt:
        # sheet.delete_rows(2)
        print(mpr, object, control_time, group, employee, sep='\t')


sheet = book['ЗНР']

for row in range(2, sheet.max_row + 1): 
    znr = sheet[row][1].value
    object = sheet[row][3].value
    control_time = sheet[row][4].value
    group = sheet[row][5].value
    employee = sheet[row][6].value
    date = control_time
    date = date.replace(tzinfo=pytz.timezone('Europe/Moscow')) 
    # format = "%y/%m/%d %H:%M:%S"

    dt = datetime.now(pytz.timezone('Europe/Moscow'))
    dt = dt.replace(microsecond=0) 

    if date < dt:
        # sheet.delete_rows(2)
        print(znr, object, control_time, group, employee, sep='\t')
        

                